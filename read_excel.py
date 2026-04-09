import openpyxl
import warnings
warnings.filterwarnings('ignore')

wb = openpyxl.load_workbook(r'C:\Users\DICT\Desktop\RULES\migration rules.xlsx')

for ws in wb.worksheets:
    print(f'\n{"="*80}')
    print(f'SHEET: {ws.title}')
    print(f'{"="*80}')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        print(row)
