"""
Template Generator — Generate Excel templates with headers, dropdowns, and data validation.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config.schema import SHEET_SCHEMAS, FieldType


class TemplateGenerator:
    """Generates Excel templates with proper headers, dropdowns, and data validation."""

    # Color scheme
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    CONDITIONAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self, output_dir: str = "outputs"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def generate_all_templates(self) -> list[str]:
        """Generate templates for all 4 sheets. Returns list of file paths."""
        paths = []
        for sheet_name in SHEET_SCHEMAS:
            path = self.generate_template(sheet_name)
            paths.append(path)
        return paths

    def generate_template(self, sheet_name: str, output_path: str | None = None) -> str:
        """Generate a single sheet template."""
        if sheet_name not in SHEET_SCHEMAS:
            raise ValueError(f"Unknown sheet: {sheet_name}")

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel max sheet name length

        schema = SHEET_SCHEMAS[sheet_name]
        fields = list(schema.values())

        # --- Row 1: Headers ---
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field.name)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.THIN_BORDER

            # Color-code by required level
            if field.required.value == "YES":
                cell.fill = self.HEADER_FILL
            elif field.required.value == "CONDITIONAL":
                pass  # keep blue header

        # --- Row 2: Descriptions ---
        desc_font = Font(name="Calibri", italic=True, size=9, color="666666")
        desc_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=2, column=col_idx, value=field.description)
            cell.font = desc_font
            cell.fill = desc_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = self.THIN_BORDER

        # --- Row 3: Required level ---
        req_font = Font(name="Calibri", bold=True, size=9)
        for col_idx, field in enumerate(fields, 1):
            label = field.required.value
            cell = ws.cell(row=3, column=col_idx, value=label)
            cell.font = req_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN_BORDER
            if label == "YES":
                cell.fill = self.REQUIRED_FILL
            elif label == "CONDITIONAL":
                cell.fill = self.CONDITIONAL_FILL

        # --- Row 4+: Example row (blank but with data validation) ---
        self._add_data_validations(ws, fields, start_row=4)

        # Set column widths
        for col_idx, field in enumerate(fields, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 22

        # Freeze panes
        ws.freeze_panes = "A4"

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}1"

        # Output
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(self.output_dir, f"{sheet_name}_TEMPLATE_{timestamp}.xlsx")

        wb.save(output_path)
        return output_path

    def _add_data_validations(self, ws, fields: list, start_row: int = 4):
        """Add Excel data validations for dropdowns and input constraints."""
        for col_idx, field in enumerate(fields, 1):
            col_letter = get_column_letter(col_idx)

            if field.field_type == FieldType.ENUM and field.enum_values:
                # Dropdown validation
                enum_str = ",".join(field.enum_values)
                dv = DataValidation(
                    type="list",
                    formula1=f'"{enum_str}"',
                    allow_blank=field.required.value != "YES",
                )
                dv.error = f"Please select: {', '.join(field.enum_values)}"
                dv.errorTitle = f"Invalid {field.name}"
                dv.prompt = f"Select from: {', '.join(field.enum_values)}"
                dv.promptTitle = field.name
                range_str = f"{col_letter}{start_row}:{col_letter}1048576"
                dv.add(range_str)
                ws.add_data_validation(dv)

            elif field.field_type == FieldType.BOOLEAN:
                dv = DataValidation(type="list", formula1='"1,0"', allow_blank=True)
                dv.error = "Enter 1 (true) or 0 (false)"
                dv.errorTitle = f"Invalid {field.name}"
                range_str = f"{col_letter}{start_row}:{col_letter}1048576"
                dv.add(range_str)
                ws.add_data_validation(dv)

            elif field.field_type == FieldType.DATE:
                dv = DataValidation(type="date", operator="greaterThanOrEqual", formula1='"1900-01-01"', allow_blank=True)
                dv.error = "Enter a valid date"
                dv.errorTitle = f"Invalid {field.name}"
                range_str = f"{col_letter}{start_row}:{col_letter}1048576"
                dv.add(range_str)
                ws.add_data_validation(dv)

            elif field.field_type == FieldType.NUMBER or field.field_type == FieldType.INTEGER:
                dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
                dv.error = "Enter a non-negative number"
                dv.errorTitle = f"Invalid {field.name}"
                range_str = f"{col_letter}{start_row}:{col_letter}1048576"
                dv.add(range_str)
                ws.add_data_validation(dv)
